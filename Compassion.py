# BISMILLAH
# THE WHITE WOLF
# NO DREAM IS TOO BIG AND NO DREAMER IS TOO SMALL


from tkinter import *
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from datetime import datetime
import time
from playsound import playsound
import threading
from win10toast import ToastNotifier

run = False

# counting total time passed
curr_time = 0
curr_min = 0
curr_hr = 0

# counting time since last break
lst_time = 0
lst_min = 0
lst_hr = 0

# break interval time set by user(default 20 mint)
break_time = 0
break_min = 20
break_hr = 0

starting_time = (datetime.today().strftime('%I') + ':' + datetime.today().strftime('%M') + " " +
                 datetime.today().strftime('%p'))


window = Tk()
window.geometry('800x450+740+0')
window.title('FROZEN')
window.configure(bg='#333533')


# function for sending notification
def notify():
    notif = ToastNotifier()
    notif.show_toast("BREAK TIME", "Hey, It's time to take a break."
    " It's good for your health. So take a break and recharge yourself.", duration=10, icon_path=r'alphabet_96900 (1).ico')


# playing the voice message
def play():
    playsound(r'VoiceMsg.wav')


# starting the timer (main starting)
def watch():

    # function for running the time loop
    def increment():
        global curr_time, curr_min, curr_hr, lst_time, lst_min, lst_hr

        if run:
            curr_time += 1
            lst_time += 1

            # Total time from start increasing
            if curr_time >= 60:
                curr_time = 0
                curr_min += 1

                if curr_min >= 60:
                    curr_min = 0
                    curr_hr += 1

            # Total time from last break increasing
            if lst_time >= 60:
                lst_time = 0
                lst_min += 1

                if lst_min >= 60:
                    lst_min = 0
                    lst_hr += 1

            # updating the time boxes
            tot_tm['text'] = 'Time Spent: '+str(curr_hr)+' Hour '+str(curr_min)+' Minute '+str(curr_time)+' Seconds'
            brk_tm['text'] = 'Since Last Break: '+str(lst_hr)+' Hour '+str(lst_min)+' Minute '+str(lst_time)+' Seconds'
            # window.after(1000, increment)

            # Checking for the break
            if lst_time >= break_time and lst_min >= break_min and lst_hr >= break_hr:
                lst_time = 0
                lst_min = 0
                lst_hr = 0

                # sending notification and playing the voice message
                voice = threading.Thread(target=play, args=())
                msg = threading.Thread(target=notify, args=())
                msg.start()
                voice.start()

            window.after(1000, increment)
    increment()


# event handler for start/stop button click
def stat():
    global run, break_hr, break_min

    if btn['text'] == 'START':
        btn["text"] = "STOP"
        run = True
        break_hr = int(inp_hr.get())
        break_min = int(inp_min.get())
        watch()
    else:
        btn["text"] = "START"       # stopping the time count
        run = False


# changing the break time interval
def reset():
    global break_hr, break_min
    break_hr = int(inp_hr.get())
    break_min = int(inp_min.get())


# starting the break time count from the start
def startover():
    global lst_hr, lst_min, lst_time
    lst_hr = 0
    lst_min = 0
    lst_time = 0


def cellname(let1, num1, let2, num2):
    return let1+str(num1)+":"+let2+str(num2)

# Adding the session data to the excel file
def addtofile():

    dat = datetime.today().strftime('%d') + " " + datetime.today().strftime('%b')
    din = datetime.today().strftime('%A')
    frm = starting_time
    to = datetime.today().strftime('%I') + ':' + datetime.today().strftime('%M') + " " + datetime.today().strftime('%p')
    duration = str(curr_hr) + " hour " + str(curr_min) + " min"

    try:
        wb = load_workbook(filename="stat_box.xlsx")
    except FileNotFoundError:
        wb = Workbook()
        wb.save("stat_box.xlsx")

    st = wb.worksheets[0]
    st.title = 'Frozen'

    # styling excel file
    align_center = Alignment(vertical='center', horizontal='center')
    align_right = Alignment(vertical='center', horizontal='right')

    title_font = Font(name='Comic Sans MS', size=18, bold=True, italic=True, vertAlign=None,
                              underline='none', strike=False, color='0000FF')
    header_font = Font(name='Comic Sans MS', size=14, bold=False, italic=False, vertAlign=None,
                              underline='none', strike=False, color='FF0000')
    info_font = Font(name='Comic Sans MS', size=10, bold=False, italic=False, vertAlign=None,
                              underline='none', strike=False, color='000000')
    total_font = Font(name='Comic Sans MS', size=14, bold=False, italic=False, vertAlign=None,
                              underline='none', strike=False, color='FF0000')
    # end of styling excel file

    # At the start of the blank sheet
    if st.max_row == 1:

        # Merging 10 cells for the title
        st.merge_cells('A1:J1')

        # Merging two cells for each header
        st.merge_cells('A2:B2')
        st.merge_cells('C2:D2')
        st.merge_cells('E2:F2')
        st.merge_cells('G2:H2')
        st.merge_cells('I2:J2')

        # Merging two cells for each info
        st.merge_cells('A3:B3')
        st.merge_cells('C3:D3')
        st.merge_cells('E3:F3')
        st.merge_cells('G3:H3')
        st.merge_cells('I3:J3')

        # Merging 6 and 4 cells for total
        st.merge_cells('A4:F4')
        st.merge_cells('G4:J4')

        st['A1'].value = "WEEKDAY: " + str(int(datetime.today().strftime('%w')) + 1)
        st['A1'].alignment = align_center
        st['A1'].font = title_font

        # updating header value
        st['A2'].value = "DATE"
        st['C2'].value = "DAY"
        st['E2'].value = "FROM"
        st['G2'].value = "TO"
        st['I2'].value = "DURATION"

        # Applying the styles to headers
        for cell in st[2]:
            cell.alignment = align_center
            cell.font = header_font

        # updating info value
        st['A3'].value = dat
        st['C3'].value = din
        st['E3'].value = frm
        st['G3'].value = to
        st['I3'].value = duration

        # Applying the styles to info
        for cell in st[3]:
            cell.alignment = align_center
            cell.font = info_font

        # updating total value and styling
        st['A4'].value = 'Total: '
        st['G4'].value = duration

        st['A4'].font = total_font
        st['G4'].font = total_font

        st['A4'].alignment = align_right

    else:                               # updating existing sheet
        lst_r = st.max_row
        prev_dat = 'A' + str(lst_r - 1)

        if dat == st[prev_dat].value:  # updating data of the same day

            prev_tot = st['G' + str(lst_r)].value
            upd_r = str(lst_r)
            add_r = str(lst_r + 1)

            # calculating total value
            upd_hr = int(prev_tot[0]) + int(duration[0])
            upd_min = int(prev_tot[7]) + int(duration[7])
            if upd_min >= 60:
                upd_min -= 60
                upd_hr += 1

            # Emptying cells previously filled with data
            st['A' + upd_r].value = ''
            st['G' + upd_r].value = ''

            # unmerging cells which are previously merged
            st.unmerge_cells(cellname('A', upd_r, 'F', upd_r))
            st.unmerge_cells(cellname('G', upd_r, 'J', upd_r))

            # Merging two cells for each info
            st.merge_cells(cellname('A', upd_r, 'B', upd_r))
            st.merge_cells(cellname('C', upd_r, 'D', upd_r))
            st.merge_cells(cellname('E', upd_r, 'F', upd_r))
            st.merge_cells(cellname('G', upd_r, 'H', upd_r))
            st.merge_cells(cellname('I', upd_r, 'J', upd_r))

            # Merging 6 and 4 cells for total
            st.merge_cells(cellname('A', add_r, 'F', add_r))
            st.merge_cells(cellname('G', add_r, 'J', add_r))

            # updating info value
            st['A' + upd_r].value = dat
            st['C' + upd_r].value = din
            st['E' + upd_r].value = frm
            st['G' + upd_r].value = to
            st['I' + upd_r].value = duration

            # Applying the styles to info
            for cell in st[upd_r]:
                cell.alignment = align_center
                cell.font = info_font

            # updating total value and styling
            st['A' + add_r].value = 'Total: '
            st['G' + add_r].value = str(upd_hr) + " hour " + str(upd_min) + " min"

            st['A' + add_r].font = total_font
            st['G' + add_r].font = total_font

            st['A' + add_r].alignment = align_right

        else:               # creating a new day
            add_daytitle = str(lst_r + 2)
            add_dayheader = str(lst_r + 3)
            add_dayinfo = str(lst_r + 4)
            add_daytotal = str(lst_r + 5)

            # Merging 10 cells for the title
            st.merge_cells(cellname('A', add_daytitle, 'J', add_daytitle))

            # Merging two cells for each header
            st.merge_cells(cellname('A', add_dayheader, 'B', add_dayheader))
            st.merge_cells(cellname('C', add_dayheader, 'D', add_dayheader))
            st.merge_cells(cellname('E', add_dayheader, 'F', add_dayheader))
            st.merge_cells(cellname('G', add_dayheader, 'H', add_dayheader))
            st.merge_cells(cellname('I', add_dayheader, 'J', add_dayheader))

            # Merging two cells for each info
            st.merge_cells(cellname('A', add_dayinfo, 'B', add_dayinfo))
            st.merge_cells(cellname('C', add_dayinfo, 'D', add_dayinfo))
            st.merge_cells(cellname('E', add_dayinfo, 'F', add_dayinfo))
            st.merge_cells(cellname('G', add_dayinfo, 'H', add_dayinfo))
            st.merge_cells(cellname('I', add_dayinfo, 'J', add_dayinfo))

            # Merging 6 and 4 cells for total
            st.merge_cells(cellname('A', add_daytotal, 'F', add_daytotal))
            st.merge_cells(cellname('G', add_daytotal, 'J', add_daytotal))

            st['A' + add_daytitle].value = "WEEKDAY: " + str(int(datetime.today().strftime('%w')) + 1)
            st['A' + add_daytitle].alignment = align_center
            st['A' + add_daytitle].font = title_font

            # updating header value
            st['A' + add_dayheader].value = "DATE"
            st['C' + add_dayheader].value = "DAY"
            st['E' + add_dayheader].value = "FROM"
            st['G' + add_dayheader].value = "TO"
            st['I' + add_dayheader].value = "DURATION"

            # Applying the styles to headers
            for cell in st[add_dayheader]:
                cell.alignment = align_center
                cell.font = header_font

            # updating info value
            st['A' + add_dayinfo].value = dat
            st['C' + add_dayinfo].value = din
            st['E' + add_dayinfo].value = frm
            st['G' + add_dayinfo].value = to
            st['I' + add_dayinfo].value = duration

            # Applying the styles to info
            for cell in st[add_dayinfo]:
                cell.alignment = align_center
                cell.font = info_font

            # updating total value and styling
            st['A' + add_daytotal].value = 'Total: '
            st['G' + add_daytotal].value = duration

            st['A' + add_daytotal].font = total_font
            st['G' + add_daytotal].font = total_font

            st['A' + add_daytotal].alignment = align_right

    try:
        wb.save("stat_box.xlsx")
    except PermissionError:
        print("Error")
    sys.exit()
# End of adding the session data to the excel file



# gui formation part

# Main Title
header = Label(window, bg='#333533', text=f"CHECK YOUR CURRENT USAGE STATISTICS",  fg="red", height="3",
        font=('Lucida Handwriting', "16", "bold italic"))

# the two time boxes shown
brk_tm = Label(window, bg='#333533', text='Since Last Break: 0 Hour 0 Minute 0 Second', fg='#00a8e8', font=('Comic Sans MS', '14'))
tot_tm = Label(window, bg='#333533', text='Time Spent: 0 Hour 0 Minute 0 Second', height='1', fg='#00a8e8', font=('Comic Sans MS', '14'))

# taking entry/input for break time interval from user
inp_frame = Frame(window, bg='#333533',  pady='2')
inp_lab = Label(inp_frame, bg='#333533', text="Desired Break Time: ", fg='#00a8e8', height='1', font=('Comic Sans MS', '14'))
inp_hr = Entry(inp_frame, bg='#FFFFFF', width='10', bd='3', font=('Comic Sans MS', '10', 'bold'))
hr_unit = Label(inp_frame, bg='#333533', text="Hour", fg='#00a8e8', font=('Comic Sans MS', '14'))
inp_min = Entry(inp_frame, bg='#FFFFFF', width='10', bd='3', font=('Comic Sans MS', '10', 'bold'))
min_unit = Label(inp_frame, bg='#333533', text="Minutes", fg='#00a8e8', font=('Comic Sans MS', '14'))

# Reset, startover, start/stop, close buttons
btn_frame = Frame(window, bg='#333533')
reset_interval = Button(btn_frame, bg='#93e1d8', text="RESET BREAK INTERVAL", cursor='hand2', bd='5', activebackground='black',
            activeforeground='white', font=('Comic Sans MS', '8'), command=reset)
start_over = Button(btn_frame, bg='#93e1d8', text="STARTOVER BREAK TIME", cursor='hand2', bd='5', activebackground='black',
            activeforeground='white', font=('Comic Sans MS', '8'), command=startover)

btn = Button(window, bg='#93e1d8', text="START", cursor='hand2', bd='5', width='7', activebackground='black',
            activeforeground='white', font=('Comic Sans MS', '12'), command=stat)
record_btn = Button(window, bg='#93e1d8', text="RECORD AND CLOSE", cursor='hand2', bd='5', width='18', activebackground='black',
            activeforeground='white', font=('Comic Sans MS', '12'), command=addtofile)


# packing section
header.pack()
brk_tm.pack()
tot_tm.pack()
inp_frame.pack(pady=(30, 5))
inp_lab.grid(row='0', column='0')
inp_hr.grid(row='0', column='1')
hr_unit.grid(row='0', column='2', padx=(2, 15))
inp_min.grid(row='0', column='3')
min_unit.grid(row='0', column='4')
btn_frame.pack(pady=(5, 30))
reset_interval.grid(row='0', column='0', padx=(5, 15))
start_over.grid(row='0', column='1', padx=(15, 5))
btn.pack()
record_btn.pack(pady=(40, 5))

# inserting the default value
inp_hr.insert(4, 0)
inp_min.insert(4, 20)
inp_hr.focus_set()

print("m")
window.mainloop()